from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


BRISBANE = ZoneInfo("Australia/Brisbane")
MONEY = Decimal("0.01")


def clean_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def key_text(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def decimal_value(value) -> Decimal:
    raw = clean_text(value).replace("$", "").replace(",", "")
    if not raw:
        return Decimal("0")
    try:
        return Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid numeric value: {value!r}") from exc


def money(value) -> Decimal:
    return decimal_value(value).quantize(MONEY, rounding=ROUND_HALF_UP)


def money_text(value) -> str:
    return format(money(value), ".2f")


def numeric_text(value, places: int = 6) -> str:
    quantizer = Decimal(1).scaleb(-places)
    return format(decimal_value(value).quantize(quantizer, rounding=ROUND_HALF_UP), f".{places}f")


def invoice_number(value) -> int | None:
    raw = clean_text(value)
    if not raw or not re.fullmatch(r"\d+", raw):
        return None
    return int(raw)


def parse_datetime(value) -> datetime | None:
    raw = clean_text(value)
    if not raw:
        return None
    formats = (
        "%b %d, %Y (%I:%M %p)",
        "%d %b, %Y (%I:%M %p)",
        "%b %d, %Y",
        "%d %b, %Y",
        "%d/%m/%Y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=BRISBANE)
        except ValueError:
            continue
    raise ValueError(f"Unsupported RepairDesk date: {value!r}")


def iso_timestamp(value: datetime) -> str:
    return value.isoformat(timespec="seconds")


def normalize_phone(value) -> str:
    digits = re.sub(r"\D", "", clean_text(value))
    if digits.startswith("61") and len(digits) == 11:
        digits = "0" + digits[2:]
    return digits


def normalize_email(value) -> str:
    email = clean_text(value).lower()
    return email if re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email) else ""


def load_invoice_export(path: Path, store_name: str, min_invoice: int, max_invoice: int):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    required = {
        "Invoice Date", "Invoice Paid date", "Invoice number", "Store name", "Ticket number",
        "Item Code", "Item Sku", "Item Category", "Item Name", "Description", "Imei/Serial",
        "Warranty Duration", "Warranty Start Date", "Warranty End Date", "PO/SO", "Price", "Tax",
        "Quantity", "Sub Total", "Amount Paid", "Due Amount", "Payment Method", "Customer First Name",
        "Customer Last Name", "Customer Email", "Customer Phone", "Customer Mobile", "Customer Address",
        "Customer City", "Customer State", "Customer Postcode", "Customer Country", "Customer Driving License",
        "Invoice Notes",
    }
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"Invoice export is missing columns: {', '.join(missing)}")

    selected = []
    for values in rows:
        row = dict(zip(headers, values))
        number = invoice_number(row["Invoice number"])
        if key_text(row["Store name"]) != key_text(store_name):
            continue
        if number is None or not min_invoice <= number <= max_invoice:
            continue
        row["_invoice_number"] = number
        selected.append(row)
    workbook.close()
    return headers, selected


def load_item_report(path: Path, store_name: str, min_invoice: int, max_invoice: int):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = [clean_text(value) for value in (reader.fieldnames or [])]
        required = {
            "Store Name", "Date", "Invoice ID", "Customer", "Email Address", "Phone Number", "Type",
            "Item ID", "SKU", "Ticket ID", "Created By", "Assigned To", "Category", "Manufacturer",
            "Device", "Product Name", "Serial", "Color", "Size", "Network", "Condition", "Quantity",
            "Total Sales", "Discount", "COGS", "Tax",
        }
        missing = sorted(required.difference(headers))
        if missing:
            raise ValueError(f"Item report is missing columns: {', '.join(missing)}")

        selected = []
        for raw_row in reader:
            row = {clean_text(key): value for key, value in raw_row.items() if key is not None}
            number = invoice_number(row.get("Invoice ID"))
            if key_text(row.get("Store Name")) != key_text(store_name):
                continue
            if number is None or not min_invoice <= number <= max_invoice:
                continue
            row["_invoice_number"] = number
            selected.append(row)
    return headers, selected


def workbook_line_key(row) -> tuple:
    return (
        clean_text(row.get("Item Code")),
        clean_text(row.get("Item Sku")),
        key_text(row.get("Item Name")),
        money_text(row.get("Quantity")),
        money_text(row.get("Sub Total")),
    )


def report_line_key(row) -> tuple:
    gross = money(decimal_value(row.get("Total Sales")) + decimal_value(row.get("Tax")))
    return (
        clean_text(row.get("Item ID")),
        clean_text(row.get("SKU")),
        key_text(row.get("Product Name")),
        money_text(row.get("Quantity")),
        money_text(gross),
    )


def item_identity_key(item_id, sku) -> tuple:
    return (
        clean_text(item_id),
        clean_text(sku).replace("\t", ""),
    )


def choose_staff(rows) -> tuple[str, list[str]]:
    names = [
        clean_text(row.get("Created By"))
        for row in rows
        if clean_text(row.get("Created By")) not in {"", "-"}
    ]
    if not names:
        return "Unknown Staff", []
    counts = Counter(names)
    ordered = sorted(counts, key=lambda name: (-counts[name], name.lower()))
    return ordered[0], ordered


def customer_payload(workbook_rows, report_rows) -> dict:
    source = workbook_rows[0]
    first_name = clean_text(source.get("Customer First Name"))
    last_name = clean_text(source.get("Customer Last Name"))
    display_name = clean_text(f"{first_name} {last_name}")
    report_customer = next((clean_text(row.get("Customer")) for row in report_rows if clean_text(row.get("Customer")) not in {"", "-"}), "")
    if not display_name:
        display_name = report_customer
    if not display_name or key_text(display_name) in {"walkin customer", "walk in customer"}:
        display_name = "Walk-in Customer"

    phone = normalize_phone(source.get("Customer Mobile")) or normalize_phone(source.get("Customer Phone"))
    if not phone:
        phone = normalize_phone(next((row.get("Phone Number") for row in report_rows if clean_text(row.get("Phone Number"))), ""))
    email = normalize_email(source.get("Customer Email"))
    if not email:
        email = normalize_email(next((row.get("Email Address") for row in report_rows if clean_text(row.get("Email Address"))), ""))

    return {
        "name": display_name,
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone,
        "email": email,
        "address": clean_text(source.get("Customer Address")),
        "city": clean_text(source.get("Customer City")),
        "state": clean_text(source.get("Customer State")),
        "postcode": clean_text(source.get("Customer Postcode")),
        "country": clean_text(source.get("Customer Country")),
        "driving_licence": clean_text(source.get("Customer Driving License")),
    }


def payment_payload(workbook_rows, invoice_at: datetime) -> list[dict]:
    grouped = defaultdict(Decimal)
    payment_times = {}
    for row in workbook_rows:
        amount = money(row.get("Amount Paid"))
        if amount == 0:
            continue
        method = clean_text(row.get("Payment Method"))
        if method in {"", "-"}:
            method = "Unknown"
        paid_at = parse_datetime(row.get("Invoice Paid date")) or invoice_at
        key = (method, iso_timestamp(paid_at))
        grouped[key] += amount
        payment_times[key] = paid_at
    return [
        {
            "method": method,
            "amount": money_text(amount),
            "taken_at": iso_timestamp(payment_times[(method, paid_at)]),
        }
        for (method, paid_at), amount in sorted(grouped.items(), key=lambda item: (item[0][1], item[0][0].lower()))
    ]


def line_payload(workbook_row, report_row, invoice_notes: str, line_number: int) -> dict:
    quantity_value = decimal_value(workbook_row.get("Quantity"))
    if quantity_value != quantity_value.to_integral_value() or quantity_value == 0:
        raise ValueError(
            f"Invoice {workbook_row['_invoice_number']} line {line_number} has unsupported quantity {quantity_value}"
        )
    quantity = int(quantity_value)
    line_total = money(workbook_row.get("Sub Total"))
    tax = money(workbook_row.get("Tax"))
    ex_gst = money(line_total - tax)
    unit_price = line_total / Decimal(quantity)
    item_id = clean_text(workbook_row.get("Item Code"))
    sku = clean_text(workbook_row.get("Item Sku"))
    description = clean_text(workbook_row.get("Description"))
    note = description
    report_row = report_row or {}

    return {
            "line_number": line_number,
            "line_type": "retail",
            "legacy_import": True,
        "product_id": f"repairdesk:{item_id or sku or workbook_row['_invoice_number']}",
        "sku": sku,
        "name": clean_text(workbook_row.get("Item Name")) or "RepairDesk item",
        "category": clean_text(workbook_row.get("Item Category")),
        "quantity": quantity,
        "unit_price": numeric_text(unit_price),
        "line_total": money_text(line_total),
        "line_payload": {
            "legacy_import": True,
            "source_system": "repairdesk",
            "source_item_id": item_id,
            "source_ticket_id": clean_text(workbook_row.get("Ticket number")) or clean_text(report_row.get("Ticket ID")),
            "source_type": clean_text(report_row.get("Type")),
            "source_assigned_to": clean_text(report_row.get("Assigned To")),
            "source_manufacturer": clean_text(report_row.get("Manufacturer")),
            "source_device": clean_text(report_row.get("Device")),
            "source_serial": clean_text(workbook_row.get("Imei/Serial")) or clean_text(report_row.get("Serial")),
            "source_color": clean_text(report_row.get("Color")),
            "source_size": clean_text(report_row.get("Size")),
            "source_network": clean_text(report_row.get("Network")),
            "source_condition": clean_text(report_row.get("Condition")),
            "source_total_sales_ex_gst": money_text(ex_gst),
            "source_tax": money_text(tax),
            "source_discount": money_text(report_row.get("Discount")),
            "source_cogs": money_text(report_row.get("COGS")),
            "source_description": description,
            "source_invoice_notes": invoice_notes,
            "source_warranty_duration": clean_text(workbook_row.get("Warranty Duration")),
            "source_warranty_start_date": clean_text(workbook_row.get("Warranty Start Date")),
            "source_warranty_end_date": clean_text(workbook_row.get("Warranty End Date")),
            "source_po_so": clean_text(workbook_row.get("PO/SO")),
            "note": note,
        },
    }


def sql_json(value) -> str:
    serialized = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return "'" + serialized.replace("'", "''") + "'::jsonb"


def build_import(args) -> dict:
    if args.min_invoice < 1 or args.max_invoice < args.min_invoice:
        raise ValueError("Invoice range must be positive and ordered")

    invoice_headers, workbook_rows = load_invoice_export(
        args.invoice_export, args.store_name, args.min_invoice, args.max_invoice
    )
    report_headers, report_rows = load_item_report(
        args.item_report, args.store_name, args.min_invoice, args.max_invoice
    )

    workbook_by_invoice = defaultdict(list)
    for row in workbook_rows:
        workbook_by_invoice[row["_invoice_number"]].append(row)
    report_by_invoice = defaultdict(list)
    for row in report_rows:
        report_by_invoice[row["_invoice_number"]].append(row)

    workbook_ids = set(workbook_by_invoice)
    report_ids = set(report_by_invoice)
    missing_in_workbook = sorted(report_ids - workbook_ids)
    missing_in_report = sorted(workbook_ids - report_ids)

    total_mismatches = []
    line_match_failures = []
    multiple_staff = []
    prepared = []
    method_counts = Counter()
    staff_counts = Counter()
    invoices_with_description = 0
    invoices_with_notes = 0

    for number in sorted(workbook_ids):
        report_group = report_by_invoice.get(number, [])
        workbook_group = workbook_by_invoice[number]
        workbook_total = sum((money(row.get("Sub Total")) for row in workbook_group), Decimal("0"))

        report_queues = defaultdict(list)
        for report_row in report_group:
            report_queues[item_identity_key(report_row.get("Item ID"), report_row.get("SKU"))].append(report_row)

        matched_report_rows = []
        for workbook_row in workbook_group:
            key = item_identity_key(workbook_row.get("Item Code"), workbook_row.get("Item Sku"))
            matched = report_queues[key].pop(0) if report_queues[key] else None
            if matched is None:
                line_match_failures.append({
                    "invoice_number": number,
                    "item_id": clean_text(workbook_row.get("Item Code")),
                    "sku": clean_text(workbook_row.get("Item Sku")),
                    "name": clean_text(workbook_row.get("Item Name")),
                    "quantity": money_text(workbook_row.get("Quantity")),
                    "gross_total": money_text(workbook_row.get("Sub Total")),
                })
            matched_report_rows.append(matched)

        invoice_at_values = [parse_datetime(row.get("Invoice Date")) for row in workbook_group]
        invoice_at = next((value for value in invoice_at_values if value is not None), None)
        if invoice_at is None:
            invoice_at = parse_datetime(report_group[0].get("Date"))
        if invoice_at is None:
            raise ValueError(f"Invoice {number} does not have a date")

        staff_name, staff_names = choose_staff(report_group)
        if len(staff_names) > 1:
            multiple_staff.append({"invoice_number": number, "staff_names": staff_names})
        staff_counts[staff_name] += 1

        customer = customer_payload(workbook_group, report_group)
        invoice_notes = next((clean_text(row.get("Invoice Notes")) for row in workbook_group if clean_text(row.get("Invoice Notes"))), "")
        if invoice_notes:
            invoices_with_notes += 1
        if any(clean_text(row.get("Description")) for row in workbook_group):
            invoices_with_description += 1

        lines = [
            line_payload(workbook_row, matched_report_rows[index], invoice_notes, index + 1)
            for index, workbook_row in enumerate(workbook_group)
        ]
        total = sum((decimal_value(line["line_total"]) for line in lines), Decimal("0"))
        payments = payment_payload(workbook_group, invoice_at)
        for payment in payments:
            method_counts[payment["method"]] += 1
        amount_paid = sum((decimal_value(payment["amount"]) for payment in payments), Decimal("0"))
        payment_methods = sorted({payment["method"] for payment in payments})
        summary_method = payment_methods[0] if len(payment_methods) == 1 else "Split" if payment_methods else "Unpaid"
        payment_status = "paid" if amount_paid + Decimal("0.01") >= total else "deposit"

        raw_payload_for_hash = {
            "invoice_number": number,
            "invoice_at": iso_timestamp(invoice_at),
            "staff_name": staff_name,
            "customer": customer,
            "invoice_notes": invoice_notes,
            "total": money_text(total),
            "amount_paid": money_text(amount_paid),
            "payments": payments,
            "items": lines,
        }
        source_hash = hashlib.sha256(
            json.dumps(raw_payload_for_hash, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()

        prepared.append({
            "source_system": "repairdesk",
            "source_store_name": args.store_name,
            "source_hash": source_hash,
            "order_code": f"RD-TW-INV-{number}",
            "invoice_number": number,
            "business_date": invoice_at.date().isoformat(),
            "created_at": iso_timestamp(invoice_at),
            "staff_name": staff_name,
            "customer_name": customer["name"],
            "customer_phone": customer["phone"],
            "customer_email": customer["email"],
            "payment_method": summary_method,
            "total": money_text(total),
            "amount_paid": money_text(amount_paid),
            "payment_status": payment_status,
            "order_payload": {
                "legacy_import": True,
                "source_system": "repairdesk",
                "source_store_name": args.store_name,
                "source_invoice_number": number,
                "source_hash": source_hash,
                "source_invoice_notes": invoice_notes,
                "source_customer": customer,
                "source_import_file": args.invoice_export.name,
                "source_item_report_file": args.item_report.name,
            },
            "items": lines,
            "payments": payments,
        })

    all_expected = set(range(args.min_invoice, args.max_invoice + 1))
    missing_invoice_numbers = sorted(all_expected - report_ids)
    imported_total = sum((decimal_value(invoice["total"]) for invoice in prepared), Decimal("0"))
    imported_paid_total = sum((decimal_value(invoice["amount_paid"]) for invoice in prepared), Decimal("0"))

    blocking_issues = {
        "missing_in_invoice_export": missing_in_workbook,
        "missing_in_item_report": missing_in_report,
    }
    manifest = {
        "source": {
            "invoice_export": str(args.invoice_export),
            "item_report": str(args.item_report),
            "store_name": args.store_name,
            "invoice_range": [args.min_invoice, args.max_invoice],
            "invoice_export_headers": invoice_headers,
            "item_report_headers": report_headers,
        },
        "counts": {
            "invoice_export_rows": len(workbook_rows),
            "item_report_rows": len(report_rows),
            "invoice_export_invoices": len(workbook_ids),
            "item_report_invoices": len(report_ids),
            "prepared_invoices": len(prepared),
            "prepared_lines": sum(len(invoice["items"]) for invoice in prepared),
            "prepared_payments": sum(len(invoice["payments"]) for invoice in prepared),
            "missing_invoice_numbers": len(missing_invoice_numbers),
            "invoices_with_line_descriptions": invoices_with_description,
            "invoices_with_invoice_notes": invoices_with_notes,
            "invoices_with_multiple_staff_names": len(multiple_staff),
            "customers_with_phone": sum(bool(invoice["customer_phone"]) for invoice in prepared),
            "customers_with_email": sum(bool(invoice["customer_email"]) for invoice in prepared),
            "negative_refund_invoices": sum(decimal_value(invoice["total"]) < 0 for invoice in prepared),
            "zero_total_invoices": sum(decimal_value(invoice["total"]) == 0 for invoice in prepared),
            "deposit_or_partially_paid_invoices": sum(invoice["payment_status"] == "deposit" for invoice in prepared),
            "invoices_without_payment_rows": sum(not invoice["payments"] for invoice in prepared),
        },
        "totals": {
            "invoice_total": money_text(imported_total),
            "amount_paid": money_text(imported_paid_total),
            "balance_due": money_text(imported_total - imported_paid_total),
        },
        "missing_invoice_numbers": missing_invoice_numbers,
        "staff_invoice_counts": dict(sorted(staff_counts.items(), key=lambda item: (-item[1], item[0].lower()))),
        "payment_invoice_counts": dict(sorted(method_counts.items(), key=lambda item: (-item[1], item[0].lower()))),
        "multiple_staff_invoices": multiple_staff,
        "warnings": {
            "invoice_total_mismatches": total_mismatches,
            "line_match_failures": line_match_failures,
        },
        "blocking_issues": blocking_issues,
    }

    args.output_dir.mkdir(parents=True, exist_ok=True)
    (args.output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (args.output_dir / "prepared-invoices.json").write_text(json.dumps(prepared, ensure_ascii=False) + "\n", encoding="utf-8")

    batch_count = 0
    for offset in range(0, len(prepared), args.batch_size):
        batch_count += 1
        batch = prepared[offset:offset + args.batch_size]
        sql = f"select public.import_repairdesk_sales_batch({sql_json(batch)});\n"
        (args.output_dir / f"batch-{batch_count:03d}.sql").write_text(sql, encoding="utf-8")
    manifest["counts"]["sql_batches"] = batch_count
    (args.output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if any(blocking_issues.values()):
        print(json.dumps(manifest, ensure_ascii=False, indent=2))
        raise RuntimeError("RepairDesk sources do not reconcile; import SQL was generated for inspection only")
    return manifest


def parse_args():
    parser = argparse.ArgumentParser(description="Prepare idempotent RepairDesk Toowong invoice import batches.")
    parser.add_argument("--invoice-export", type=Path, default=Path.home() / "Downloads" / "Invoices2026-08-20-23-08-59.xlsx")
    parser.add_argument("--item-report", type=Path, default=Path.home() / "Downloads" / "Item Wise Sales Report.csv")
    parser.add_argument("--output-dir", type=Path, default=Path(".codex-temp/repairdesk-toowong-invoice-import"))
    parser.add_argument("--store-name", default="TechM8 Toowong")
    parser.add_argument("--min-invoice", type=int, default=1)
    parser.add_argument("--max-invoice", type=int, default=3848)
    parser.add_argument("--batch-size", type=int, default=50)
    return parser.parse_args()


if __name__ == "__main__":
    try:
        result = build_import(parse_args())
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise
